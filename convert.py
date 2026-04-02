"""
convert.py — OAV Report Card Data Converter
============================================
Reads template.xlsx and outputs one JSON file per class into data/

Usage:
    python3 convert.py                     # uses template.xlsx in same folder
    python3 convert.py path/to/file.xlsx   # custom file path

Template format (one sheet per class, named "Class 6" … "Class 12"):
    Row 1  : Headers (fixed, do not edit)
    Row 2  : Max marks (fixed, do not edit)
    Row 3+ : Student data (one student per row)

See template.xlsx for the exact column layout.
"""

import sys
import json
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────────────────────

# Map sheet names → output JSON filenames (without .json)
SHEET_TO_CLASS = {
    "Class 6":  "class6",
    "Class 7":  "class7",
    "Class 8":  "class8",
    "Class 9":  "class9",
    "Class 10": "class10",
    "Class 11": "class11",
    "Class 12": "class12",
}

SUBJECTS = ["ENGLISH", "ODIA", "HINDI", "SANSKRIT", "MATHEMATICS", "SCIENCE", "SOCIAL SCIENCE", "ICT"]

# Exam key in template header → JSON key, full marks
EXAM_MAP = {
    "PT1": ("PT-I",   40),
    "PT2": ("PT-II",  40),
    "HFY": ("HFY",   100),
    "PT3": ("PT-III", 40),
    "PT4": ("PT-IV",  40),
    "ANN": ("ANNUAL", 80),
    "IA":  ("IA",     20),
}

CO_SCHO_MAP = {
    "Work Education":  "WORK EDUCATION",
    "Art Education":   "ART EDUCATION",
    "Health PE":       "HEALTH & PHYSICAL EDUCATION",
    "Discipline":      "DISCIPLINE",
    "Sports":          "SPORTS",
}

GRADE_SCALE = [
    (91, "A1"), (81, "A2"), (71, "B1"), (61, "B2"),
    (51, "C1"), (41, "C2"), (33, "D"),  (0,  "E"),
]

OUTPUT_DIR = Path(__file__).parent / "data"

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_grade(pct):
    for threshold, grade in GRADE_SCALE:
        if pct >= threshold:
            return grade
    return "E"

def safe_val(cell_value):
    """Return int/float, 'ABS' for absent, or None for blank."""
    if cell_value is None or str(cell_value).strip() == "":
        return None
    s = str(cell_value).strip().upper()
    if s == "ABS":
        return "ABS"
    try:
        f = float(cell_value)
        return int(f) if f == int(f) else round(f, 2)
    except (ValueError, TypeError):
        return str(cell_value).strip()

def safe_str(cell_value):
    if cell_value is None:
        return None
    return str(cell_value).strip() or None

def build_column_map(header_row):
    """
    Returns a dict mapping header string → 0-based column index.
    Skips None/empty cells.
    """
    col_map = {}
    for idx, cell in enumerate(header_row):
        val = safe_str(cell.value)
        if val:
            col_map[val.upper()] = idx
    return col_map

# ── Main processing ───────────────────────────────────────────────────────────

def process_sheet(ws, sheet_name):
    rows = list(ws.iter_rows())
    if len(rows) < 3:
        print(f"  [SKIP] {sheet_name}: fewer than 3 rows, nothing to process")
        return {}

    header_row = rows[0]   # Row 1
    # Row 2 is max-marks reference — skip
    data_rows  = rows[2:]  # Row 3+

    col_map = build_column_map(header_row)

    # Helper: get column index by header key (case-insensitive)
    def col(key):
        return col_map.get(key.upper())

    # Verify essential columns exist
    for req in ["ADMISSION NO", "NAME", "ROLL NO", "SECTION"]:
        if col(req) is None:
            print(f"  [ERROR] {sheet_name}: required column '{req}' not found in header row")
            print(f"          Found headers: {list(col_map.keys())[:10]}...")
            return {}

    students = []

    for row in data_rows:
        vals = [c.value for c in row]

        adm  = safe_str(vals[col("ADMISSION NO")]) if col("ADMISSION NO") is not None else None
        name = safe_str(vals[col("NAME")])          if col("NAME") is not None else None

        # Skip blank rows
        if not adm and not name:
            continue
        if not adm:
            adm = f"UNKNOWN_{name}"
        if not name:
            print(f"  [WARN] Row with admission {adm} has no name — skipping")
            continue

        student = {
            "name":        name,
            "father_name": safe_str(vals[col("FATHER NAME")]) if col("FATHER NAME") is not None else None,
            "dob":         safe_str(vals[col("DOB")])          if col("DOB") is not None else None,
            "roll":        safe_val(vals[col("ROLL NO")])       if col("ROLL NO") is not None else None,
            "section":     safe_str(vals[col("SECTION")])       if col("SECTION") is not None else None,
            "attendance":  safe_val(vals[col("ATTENDANCE")])    if col("ATTENDANCE") is not None else None,
        }

        # Co-scholastic
        co = {}
        for tmpl_key, json_key in CO_SCHO_MAP.items():
            c_idx = col(tmpl_key)
            if c_idx is not None:
                co[json_key] = safe_str(vals[c_idx])
        student["co_scholastic"] = co

        # Exam marks
        exams = {}
        for tmpl_prefix, (json_key, max_marks) in EXAM_MAP.items():
            subj_marks = {}
            for sub in SUBJECTS:
                hdr = f"{tmpl_prefix} {sub}".upper()
                c_idx = col(hdr)
                if c_idx is not None:
                    subj_marks[sub] = safe_val(vals[c_idx])
                else:
                    subj_marks[sub] = None
            exams[json_key] = subj_marks

        # Compute TOTAL per subject (ANN + IA)
        totals = {}
        for sub in SUBJECTS:
            ann = exams["ANNUAL"].get(sub)
            ia  = exams["IA"].get(sub)
            if ann == "ABS" or ia == "ABS":
                totals[sub] = "ABS"
            elif isinstance(ann, (int, float)) and isinstance(ia, (int, float)):
                totals[sub] = ann + ia
            else:
                totals[sub] = None
        exams["TOTAL"] = totals

        student["exams"] = exams

        # Overall total and percentage (from TOTAL, skip ABS/None)
        total_marks = 0
        total_max   = 0
        for sub in SUBJECTS:
            t = totals.get(sub)
            if isinstance(t, (int, float)):
                total_marks += t
                total_max   += 100

        student["total"]   = total_marks
        student["percent"] = round(total_marks / total_max * 100, 2) if total_max > 0 else 0
        student["grade"]   = get_grade(student["percent"])

        students.append((adm, student))

    # Assign ranks (by percent descending)
    students.sort(key=lambda x: x[1]["percent"], reverse=True)
    for rank, (adm, s) in enumerate(students, start=1):
        s["rank"] = rank

    result = {adm: s for adm, s in students}
    print(f"  [OK] {sheet_name}: {len(result)} students processed")
    return result

# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / "template.xlsx"

    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        print("Usage: python3 convert.py [path/to/template.xlsx]")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Sheets found: {wb.sheetnames}\n")

    OUTPUT_DIR.mkdir(exist_ok=True)
    generated = []

    for sheet_name, class_key in SHEET_TO_CLASS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [SKIP] Sheet '{sheet_name}' not found in workbook")
            continue

        ws = wb[sheet_name]
        class_data = process_sheet(ws, sheet_name)

        if class_data:
            out_path = OUTPUT_DIR / f"{class_key}.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(class_data, f, indent=2, ensure_ascii=False)
            print(f"         → {out_path}")
            generated.append(str(out_path))

    print(f"\nDone. {len(generated)} file(s) written:")
    for p in generated:
        print(f"  {p}")

    if not generated:
        print("\nNo files were generated. Check that your sheet names match:")
        for k in SHEET_TO_CLASS:
            print(f"  '{k}'")

if __name__ == "__main__":
    main()
